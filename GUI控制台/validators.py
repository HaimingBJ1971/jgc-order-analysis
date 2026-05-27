import os
import openpyxl
import sqlite3
import pandas as pd

def create_msg(level, code, message, file_path=None, suggestion=""):
    return {
        "level": level,
        "code": code,
        "message": message,
        "file": file_path or "",
        "suggestion": suggestion
    }

def scan_files(paths):
    """
    Scans paths (files or directories), resolves directories (1-level deep),
    ignores temporary ~$ files and report files inside 'output/' directories.
    """
    resolved_files = []
    for p in paths:
        p = os.path.abspath(p)
        if os.path.isdir(p):
            # Scan 1-level deep
            for item in os.listdir(p):
                full_path = os.path.join(p, item)
                if os.path.isfile(full_path):
                    ext = os.path.splitext(full_path)[1].lower()
                    is_db = ext in ['.db', '.sqlite']
                    if not item.startswith("~$") and (is_db or "output" not in full_path.lower()):
                        resolved_files.append(full_path)
        elif os.path.isfile(p):
            ext = os.path.splitext(p)[1].lower()
            is_db = ext in ['.db', '.sqlite']
            if not os.path.basename(p).startswith("~$") and (is_db or "output" not in p.lower()):
                resolved_files.append(p)
    return resolved_files

def classify_files(files):
    """
    Classifies files into 'pos_excel', 'table_csv', 'takeaway_excel', 'database', 'unrecognized'.
    Returns a dictionary: { 'pos_excel': [], 'table_csv': [], 'takeaway_excel': [], 'database': [], 'unrecognized': [] }
    """
    classified = {
        "pos_excel": [],
        "table_csv": [],
        "takeaway_excel": [],
        "database": [],
        "unrecognized": []
    }
    
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        basename = os.path.basename(f)
        
        if ext in ['.db', '.sqlite']:
            # Verify SQLite
            try:
                conn = sqlite3.connect(f)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                cursor.fetchall()
                conn.close()
                classified["database"].append(f)
            except Exception:
                classified["unrecognized"].append(f)
                
        elif ext in ['.xlsx', '.xls']:
            try:
                # Fast check using read_only=True
                wb = openpyxl.load_workbook(f, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                
                if "店内订单明细" in sheets and "商品-店内订单明细" in sheets:
                    classified["pos_excel"].append(f)
                elif "平台外卖订单明细" in sheets:
                    classified["takeaway_excel"].append(f)
                elif "店内订单明细" in basename or "pos" in basename.lower():
                    classified["pos_excel"].append(f)
                elif "外卖" in basename or "takeaway" in basename.lower():
                    classified["takeaway_excel"].append(f)
                else:
                    classified["unrecognized"].append(f)
            except Exception:
                # Fallback to name match
                if "店内订单明细" in basename or "pos" in basename.lower():
                    classified["pos_excel"].append(f)
                elif "外卖" in basename or "takeaway" in basename.lower():
                    classified["takeaway_excel"].append(f)
                else:
                    classified["unrecognized"].append(f)
                    
        elif ext == '.csv':
            if "桌探" in basename or "桌访" in basename:
                classified["table_csv"].append(f)
            else:
                # Quick read columns to verify
                try:
                    df = pd.read_csv(f, nrows=2)
                    cols = [str(c).strip() for c in df.columns]
                    if any(k in cols for k in ["订单号", "就餐人数", "语音转录", "总体满意度评分"]):
                        classified["table_csv"].append(f)
                    else:
                        classified["unrecognized"].append(f)
                except Exception:
                    classified["unrecognized"].append(f)
        else:
            classified["unrecognized"].append(f)
            
    return classified

# ── Feature-specific Pre-Validators ──────────────────────────

def validate_order_zhuofang(pos_files, csv_files, output_dir):
    messages = []
    
    # Check POS
    if not pos_files:
        messages.append(create_msg("error", "missing_pos", "缺少 POS 订单明细文件", suggestion="请选择或拖入有效的 POS 店内订单明细 Excel 文件"))
        return False, messages
        
    pos_file = pos_files[0]
    try:
        wb = openpyxl.load_workbook(pos_file, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        
        if "店内订单明细" not in sheets:
            messages.append(create_msg("error", "missing_pos_sheet", "POS文件缺少 '店内订单明细' 工作表", pos_file, "请使用标准的POS系统导出的Excel文件"))
        if "商品-店内订单明细" not in sheets:
            messages.append(create_msg("error", "missing_items_sheet", "POS文件缺少 '商品-店内订单明细' 工作表", pos_file, "请使用标准的POS系统导出的Excel文件"))
    except Exception as e:
        messages.append(create_msg("error", "corrupt_pos", f"POS文件格式损毁或无法读取: {e}", pos_file))
        
    # Check CSV
    if not csv_files:
        messages.append(create_msg("warning", "missing_csv", "未提供桌访反馈 CSV 数据，系统将执行纯订单合并分析", suggestion="拖入相应日期的桌访CSV可以进行就餐人数修正与深度匹配"))
    else:
        csv_file = csv_files[0]
        # Basic coding warning if not UTF-8
        try:
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                f.readline()
        except UnicodeDecodeError:
            messages.append(create_msg("warning", "csv_encoding", "桌访CSV编码格式不是 UTF-8-SIG，读取时程序将尝试自动转码", csv_file))
            
    # Check Output Dir
    if output_dir:
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir, exist_ok=True)
            except Exception:
                messages.append(create_msg("error", "output_dir_unwritable", "指定输出目录不存在且无法创建", suggestion="请选择一个有写权限的目录路径"))
        elif not os.access(output_dir, os.W_OK):
            messages.append(create_msg("error", "output_dir_unwritable", "指定输出目录不可写", suggestion="请选择一个有写权限的目录路径"))
            
    has_error = any(m["level"] == "error" for m in messages)
    return not has_error, messages

def validate_long_term(pos_files, db_path, output_dir):
    messages = []
    
    if not pos_files:
        messages.append(create_msg("error", "missing_pos_long", "未导入任何 POS 订单明细文件", suggestion="拖入或选择一个或多个 POS 订单明细 Excel 文件进行批量写库"))
        return False, messages
        
    # Verify each POS file sheets
    for pf in pos_files:
        try:
            wb = openpyxl.load_workbook(pf, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            if "店内订单明细" not in sheets or "商品-店内订单明细" not in sheets:
                messages.append(create_msg("error", "invalid_pos_structure", f"文件 {os.path.basename(pf)} 结构不合格", pf, "必须包含 '店内订单明细' 和 '商品-店内订单明细' 两个工作表"))
        except Exception:
            messages.append(create_msg("error", "invalid_pos_read", f"文件 {os.path.basename(pf)} 无法正常读取", pf))
            
    # DB path checks
    if not db_path:
        messages.append(create_msg("error", "missing_db", "缺少 SQLite 数据库路径", suggestion="必须指定长期订单分析 SQLite 数据库路径以保存历史数据"))
    else:
        db_dir = os.path.dirname(os.path.abspath(db_path))
        if not os.path.exists(db_dir):
            try:
                os.makedirs(db_dir, exist_ok=True)
            except Exception:
                messages.append(create_msg("error", "db_dir_unwritable", "指定数据库目录无法创建", suggestion="请修改数据库文件存放位置"))
        elif not os.access(db_dir, os.W_OK):
            messages.append(create_msg("error", "db_dir_unwritable", "指定数据库目录不可写", suggestion="请更换有写权限的目录存放数据库"))
            
    # Output Dir Check
    if output_dir:
        if not os.path.exists(output_dir) or not os.access(output_dir, os.W_OK):
            messages.append(create_msg("error", "output_dir_unwritable", "指定输出目录不可写或不存在"))
            
    has_error = any(m["level"] == "error" for m in messages)
    return not has_error, messages

def validate_period_compare(pos_files, db_path, mode, output_dir):
    messages = []
    
    # Check POS
    if not pos_files:
        messages.append(create_msg("error", "missing_pos", "缺少当前周期的 POS 订单明细文件", suggestion="请选择或拖入当前分析周期的 POS Excel 文件"))
        return False, messages
        
    pos_file = pos_files[0]
    try:
        wb = openpyxl.load_workbook(pos_file, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        if "店内订单明细" not in sheets or "商品-店内订单明细" not in sheets:
            messages.append(create_msg("error", "pos_structure", "当前周期POS文件结构不合格", pos_file))
    except Exception:
        messages.append(create_msg("error", "pos_read", "当前周期POS文件无法正常读取", pos_file))
        
    # Check DB
    if not db_path:
        messages.append(create_msg("error", "missing_db", "缺少 SQLite 数据库路径", suggestion="周期对比分析必须依赖已归档长期历史订单的 SQLite 数据库"))
    elif not os.path.exists(db_path):
        messages.append(create_msg("error", "db_not_exist", "SQLite 数据库文件不存在", db_path, "请指定已有数据的数据库文件，否则无法提取同比/环比数据"))
    else:
        # Check tables existence in DB
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [t[0] for t in cursor.fetchall()]
            conn.close()
            required = ['orders', 'groups', 'daily_overview', 'daily_order_counts']
            missing = [r for r in required if r not in tables]
            if missing:
                messages.append(create_msg("error", "db_missing_tables", f"数据库缺少必要的历史订单表: {', '.join(missing)}", db_path, "请先在“长期订单分析”功能中导入数据建立完整历史表"))
        except Exception as e:
            messages.append(create_msg("error", "db_corrupt", f"SQLite 数据库损坏或无法建立连接: {e}", db_path))
            
    # Check Mode
    if not mode:
        messages.append(create_msg("error", "missing_mode", "未选择对比周期模式 (week 或 month)", suggestion="请在参数区域选择对比周期模式"))
        
    # Output Dir Check
    if output_dir:
        if not os.path.exists(output_dir) or not os.access(output_dir, os.W_OK):
            messages.append(create_msg("error", "output_dir_unwritable", "指定输出目录不可写或不存在"))
            
    has_error = any(m["level"] == "error" for m in messages)
    return not has_error, messages

def validate_takeaway(takeaway_files, db_path, output_dir):
    messages = []
    
    if not takeaway_files:
        messages.append(create_msg("error", "missing_takeaway", "缺少平台外卖订单 Excel 文件", suggestion="请选择或拖入包含 '平台外卖订单明细' 工作表的外卖明细 Excel 文件"))
        return False, messages
        
    for tf in takeaway_files:
        try:
            wb = openpyxl.load_workbook(tf, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            if "平台外卖订单明细" not in sheets:
                messages.append(create_msg("error", "invalid_takeaway_sheet", f"文件 {os.path.basename(tf)} 缺少 '平台外卖订单明细' 工作表", tf))
        except Exception:
            messages.append(create_msg("error", "takeaway_read", f"文件 {os.path.basename(tf)} 无法正常读取", tf))
            
    if not db_path:
        messages.append(create_msg("warning", "missing_db_takeaway", "未指定 SQLite 数据库路径，外卖统计结果将只生成报告，不会进行数据库归档", suggestion="如果需要沉淀长期外卖报表，请在参数中指定 SQLite 数据库路径"))
        
    # Output Dir Check
    if output_dir:
        if not os.path.exists(output_dir) or not os.access(output_dir, os.W_OK):
            messages.append(create_msg("error", "output_dir_unwritable", "指定输出目录不可写或不存在"))
            
    has_error = any(m["level"] == "error" for m in messages)
    return not has_error, messages
