"""
Weekday half-year trend workbook.

The workbook is read-only against the long-term SQLite database. It uses the
kept dining-group daily aggregates already produced by the order merge pipeline:

- daily_overview category "整体" for dining-group revenue and people.
- daily_order_counts.统计消费团体数 for group/table count.
- daily_overview categories "大厅", "包间", "户外" for area people and per-person.
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import tempfile
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html import escape as html_escape
from pathlib import Path

os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-cache")
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


WEEKDAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
AREA_CATEGORIES = {
    "大厅": "大厅",
    "包房": "包间",
    "院子": "户外",
}
HEADER_ROW = 4
DATA_ROW = HEADER_ROW + 1
THEME_BLUE = "#1f4e78"
TEXT_DARK = "#2c3e50"
GRID_GREY = "#D3D3D3"


def _register_chinese_font() -> str:
    font_paths = [
        ("/System/Library/Fonts/STHeiti Medium.ttc", 0),
        ("/System/Library/Fonts/STHeiti Light.ttc", 0),
        ("/System/Library/Fonts/PingFang.ttc", 0),
        ("/System/Library/Fonts/Supplemental/Arial Unicode.ttf", None),
        ("/usr/share/fonts/truetype/wqy/wqy-microhei.ttc", 0),
        ("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc", 0),
        ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", 0),
    ]
    for font_path, subfont_index in font_paths:
        if not os.path.exists(font_path):
            continue
        try:
            name = "ChineseFont"
            if font_path.lower().endswith(".ttc") and subfont_index is not None:
                pdfmetrics.registerFont(TTFont(name, font_path, subfontIndex=subfont_index))
            else:
                pdfmetrics.registerFont(TTFont(name, font_path))
            return name
        except Exception:
            continue
    return "Helvetica"


CHINESE_FONT = _register_chinese_font()


def _p(text: object, style: ParagraphStyle) -> Paragraph:
    return Paragraph(html_escape(str(text)), style)


def _fmt_currency(value: object) -> str:
    if value in (None, ""):
        return "-"
    return f"¥{float(value):,.0f}"


def _fmt_number(value: object) -> str:
    if value in (None, ""):
        return "-"
    return f"{float(value):,.0f}"


def _fmt_average(value: object) -> str:
    if value in (None, ""):
        return "-"
    return f"¥{float(value):,.0f}"


def _pdf_styles() -> dict[str, ParagraphStyle]:
    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        "WeekdayNormal",
        parent=styles["Normal"],
        fontName=CHINESE_FONT,
        fontSize=9,
        leading=12,
    )
    return {
        "title": ParagraphStyle(
            "WeekdayTitle",
            parent=styles["Title"],
            fontName=CHINESE_FONT,
            fontSize=18,
            leading=22,
            textColor=colors.HexColor(THEME_BLUE),
            spaceAfter=14,
        ),
        "subtitle": ParagraphStyle(
            "WeekdaySubtitle",
            parent=styles["Heading2"],
            fontName=CHINESE_FONT,
            fontSize=12,
            leading=16,
            textColor=colors.HexColor(TEXT_DARK),
            spaceBefore=10,
            spaceAfter=7,
        ),
        "normal": normal,
        "note": ParagraphStyle(
            "WeekdayNote",
            parent=normal,
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#666666"),
        ),
        "cell_c": ParagraphStyle(
            "WeekdayCellC",
            parent=normal,
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
        ),
        "cell_l": ParagraphStyle(
            "WeekdayCellL",
            parent=normal,
            fontSize=7,
            leading=9,
            alignment=TA_LEFT,
        ),
        "cell_r": ParagraphStyle(
            "WeekdayCellR",
            parent=normal,
            fontSize=7,
            leading=9,
            alignment=TA_RIGHT,
        ),
        "header": ParagraphStyle(
            "WeekdayHeader",
            parent=normal,
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
            textColor=colors.white,
        ),
    }


@dataclass(frozen=True)
class TrendCell:
    revenue: float | None = None
    people: float | None = None

    @property
    def average(self) -> float | None:
        if self.revenue is None or self.people is None:
            return None
        return round(self.revenue / self.people, 2) if self.people > 0 else 0.0


@dataclass(frozen=True)
class WeekdayTrendConfig:
    db_path: Path
    output_dir: Path
    end_date: date
    store_name: str | None = None
    weeks: int = 26

    @property
    def latest_monday(self) -> date:
        return self.end_date - timedelta(days=self.end_date.weekday())

    @property
    def first_monday(self) -> date:
        return self.latest_monday - timedelta(weeks=self.weeks - 1)

    @property
    def date_tag(self) -> str:
        start = self.first_monday.strftime("%Y%m%d")
        end = (self.latest_monday + timedelta(days=6)).strftime("%Y%m%d")
        return f"{start}_{end}"

    @property
    def store_tag(self) -> str:
        return self.store_name or "双店合计"


def _parse_date(text: str) -> date:
    return datetime.strptime(text, "%Y-%m-%d").date()


def _connect_readonly(db_path: Path) -> sqlite3.Connection:
    if not db_path.exists():
        raise FileNotFoundError(f"SQLite 主库不存在: {db_path}")
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA query_only = TRUE")
    return conn


def _latest_daily_date(conn: sqlite3.Connection, store_name: str | None = None) -> date:
    if store_name:
        row = conn.execute(
            "SELECT MAX(date) FROM daily_overview WHERE store_name = ?",
            (store_name,),
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT MAX(date) FROM daily_overview WHERE store_name != '__legacy__'"
        ).fetchone()
    if not row or not row[0]:
        raise ValueError("daily_overview 中没有可用于生成周几趋势的数据")
    return _parse_date(str(row[0]))


def _date_range(config: WeekdayTrendConfig) -> tuple[str, str]:
    start = config.first_monday.strftime("%Y-%m-%d")
    end = (config.latest_monday + timedelta(days=6)).strftime("%Y-%m-%d")
    return start, end


def _fetch_overview(
    conn: sqlite3.Connection,
    config: WeekdayTrendConfig,
) -> dict[tuple[str, str], TrendCell]:
    start, end = _date_range(config)
    params: list[object] = [start, end]
    if config.store_name:
        store_clause = "AND store_name = ?"
        params.append(config.store_name)
    else:
        store_clause = "AND store_name != '__legacy__'"
    rows = conn.execute(
        f"""
        SELECT date, category, SUM(营业额) AS revenue, SUM(人数) AS people
        FROM daily_overview
        WHERE date BETWEEN ? AND ?
          {store_clause}
          AND sub_category = ''
          AND category IN ('整体', '大厅', '包间', '户外')
        GROUP BY date, category
        """,
        params,
    ).fetchall()
    return {
        (str(day), str(category)): TrendCell(
            revenue=float(revenue or 0),
            people=float(people or 0),
        )
        for day, category, revenue, people in rows
    }


def _fetch_group_counts(
    conn: sqlite3.Connection,
    config: WeekdayTrendConfig,
) -> dict[str, int]:
    start, end = _date_range(config)
    params: list[object] = [start, end]
    if config.store_name:
        store_clause = "AND store_name = ?"
        params.append(config.store_name)
    else:
        store_clause = "AND store_name != '__legacy__'"
    rows = conn.execute(
        f"""
        SELECT date, SUM(统计消费团体数)
        FROM daily_order_counts
        WHERE date BETWEEN ? AND ?
          {store_clause}
        GROUP BY date
        """,
        params,
    ).fetchall()
    return {str(day): int(count or 0) for day, count in rows}


def build_weekday_rows(
    conn: sqlite3.Connection,
    config: WeekdayTrendConfig,
) -> dict[str, list[list[object]]]:
    overview = _fetch_overview(conn, config)
    group_counts = _fetch_group_counts(conn, config)
    result: dict[str, list[list[object]]] = {name: [] for name in WEEKDAY_NAMES}

    for weekday_index, weekday_name in enumerate(WEEKDAY_NAMES):
        for week_offset in range(config.weeks):
            day = config.first_monday + timedelta(weeks=week_offset, days=weekday_index)
            day_text = day.strftime("%Y-%m-%d")
            overall = overview.get((day_text, "整体"))
            has_data = overall is not None or day_text in group_counts
            row: list[object] = [
                day,
                f"{day.isocalendar().year}W{day.isocalendar().week:02d}",
                weekday_name,
                "已入库" if has_data else "缺失",
                overall.revenue if overall else None,
                group_counts.get(day_text),
                overall.people if overall else None,
                overall.average if overall else None,
            ]
            for display_name, category in AREA_CATEGORIES.items():
                cell = overview.get((day_text, category))
                row.extend([
                    cell.people if cell else None,
                    cell.average if cell else None,
                ])
            result[weekday_name].append(row)
    return result


def _style_sheet(ws, max_data_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = f"A{DATA_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:N{max_data_row}"
    widths = {
        "A": 13, "B": 10, "C": 8, "D": 10, "E": 13, "F": 10, "G": 10,
        "H": 11, "I": 10, "J": 11, "K": 10, "L": 11, "M": 10, "N": 11,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=max_data_row, min_col=1, max_col=14):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == HEADER_ROW:
                cell.fill = header_fill
                cell.font = header_font

    for row in ws.iter_rows(min_row=DATA_ROW, max_row=max_data_row):
        row[0].number_format = "yyyy-mm-dd"
        for cell in (row[4], row[7], row[9], row[11], row[13]):
            cell.number_format = '#,##0.00'
        for cell in (row[5], row[6], row[8], row[10], row[12]):
            cell.number_format = '0'


def _add_line_chart(ws, title: str, columns: list[int], anchor: str, max_data_row: int) -> None:
    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.height = 8
    chart.width = 18
    chart.y_axis.majorGridlines = None
    chart.x_axis.title = "日期"
    data_start = HEADER_ROW
    categories = Reference(ws, min_col=1, min_row=DATA_ROW, max_row=max_data_row)
    for column in columns:
        data = Reference(ws, min_col=column, min_row=data_start, max_row=max_data_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)


def write_weekday_trend_workbook(config: WeekdayTrendConfig) -> Path:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    conn = _connect_readonly(config.db_path)
    try:
        rows_by_weekday = build_weekday_rows(conn, config)
    finally:
        conn.close()

    output_path = config.output_dir / f"周几半年趋势_{config.date_tag}_{config.store_tag}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "日期", "ISO周", "星期", "数据状态", "营业额", "团体数", "总人数", "整体人均",
        "大厅人数", "大厅人均", "包房人数", "包房人均", "院子人数", "院子人均",
    ]

    for weekday_name, rows in rows_by_weekday.items():
        ws = wb.create_sheet(weekday_name)
        ws.merge_cells("A1:N1")
        ws["A1"] = f"{config.store_tag} {weekday_name} 半年趋势"
        ws["A1"].font = Font(size=15, bold=True, color="1F4E78")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A2"] = (
            f"周期: {config.first_monday:%Y-%m-%d} ~ "
            f"{(config.latest_monday + timedelta(days=6)):%Y-%m-%d}; "
            "营业额为堂食分桌有效消费团体口径"
        )
        ws.append([])
        ws.append(headers)
        for row in rows:
            ws.append(row)

        max_data_row = HEADER_ROW + len(rows)
        _style_sheet(ws, max_data_row)
        _add_line_chart(ws, "营业额趋势", [5], "P2", max_data_row)
        _add_line_chart(ws, "团体数与人数趋势", [6, 7], "P20", max_data_row)
        _add_line_chart(ws, "人均趋势", [8, 10, 12, 14], "P38", max_data_row)

    wb.save(output_path)
    return output_path


def _pdf_table_style(header_rows: int = 1) -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, header_rows - 1), colors.HexColor(THEME_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, header_rows - 1), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(GRID_GREY)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ])


def _latest_summary_table(rows_by_weekday: dict[str, list[list[object]]], styles: dict[str, ParagraphStyle]) -> Table:
    headers = ["星期", "日期", "营业额", "团体数", "总人数", "整体人均", "大厅人均", "包房人均", "院子人均"]
    table_rows = [[_p(h, styles["header"]) for h in headers]]
    for weekday_name in WEEKDAY_NAMES:
        rows = rows_by_weekday[weekday_name]
        row = rows[-1] if rows else []
        table_rows.append([
            _p(weekday_name, styles["cell_c"]),
            _p(row[0].strftime("%Y-%m-%d") if row else "-", styles["cell_c"]),
            _p(_fmt_currency(row[4] if row else None), styles["cell_r"]),
            _p(_fmt_number(row[5] if row else None), styles["cell_r"]),
            _p(_fmt_number(row[6] if row else None), styles["cell_r"]),
            _p(_fmt_average(row[7] if row else None), styles["cell_r"]),
            _p(_fmt_average(row[9] if row else None), styles["cell_r"]),
            _p(_fmt_average(row[11] if row else None), styles["cell_r"]),
            _p(_fmt_average(row[13] if row else None), styles["cell_r"]),
        ])
    table = Table(
        table_rows,
        colWidths=[1.5*cm, 2.3*cm, 2.2*cm, 1.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm],
        repeatRows=1,
    )
    table.setStyle(_pdf_table_style())
    return table


def _weekday_detail_table(rows: list[list[object]], styles: dict[str, ParagraphStyle]) -> Table:
    headers = ["日期", "营业额", "团体", "人数", "整体人均", "大厅", "包房", "院子"]
    table_rows = [[_p(h, styles["header"]) for h in headers]]
    for row in rows:
        table_rows.append([
            _p(row[0].strftime("%m-%d"), styles["cell_c"]),
            _p(_fmt_currency(row[4]), styles["cell_r"]),
            _p(_fmt_number(row[5]), styles["cell_r"]),
            _p(_fmt_number(row[6]), styles["cell_r"]),
            _p(_fmt_average(row[7]), styles["cell_r"]),
            _p(_fmt_average(row[9]), styles["cell_r"]),
            _p(_fmt_average(row[11]), styles["cell_r"]),
            _p(_fmt_average(row[13]), styles["cell_r"]),
        ])
    table = Table(
        table_rows,
        colWidths=[1.8*cm, 2.3*cm, 1.3*cm, 1.3*cm, 1.9*cm, 1.7*cm, 1.7*cm, 1.7*cm],
        repeatRows=1,
    )
    table.setStyle(_pdf_table_style())
    return table


def _series(rows: list[list[object]], column_index: int) -> list[float]:
    values = []
    for row in rows:
        value = row[column_index]
        values.append(float(value) if value not in (None, "") else float("nan"))
    return values


def _plot_weekday_chart(rows: list[list[object]], weekday_name: str, output_dir: Path, store_tag: str) -> Path:
    plt.rcParams["font.sans-serif"] = ["Arial Unicode MS", "PingFang SC", "STHeiti", "Heiti SC", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False
    labels = [row[0].strftime("%m-%d") for row in rows]
    x = list(range(len(labels)))

    fig, axes = plt.subplots(3, 1, figsize=(9.5, 7.2), dpi=220, sharex=True)
    fig.suptitle(f"{store_tag} {weekday_name} 半年趋势", fontsize=13, fontweight="bold", color=TEXT_DARK)

    axes[0].plot(x, _series(rows, 4), marker="o", linewidth=2.2, markersize=4, color=THEME_BLUE, label="营业额")
    axes[0].set_ylabel("营业额")
    axes[0].legend(loc="upper left")

    axes[1].plot(x, _series(rows, 5), marker="o", linewidth=2.0, markersize=4, color="#27AE60", label="团体数")
    axes[1].plot(x, _series(rows, 6), marker="o", linewidth=2.0, markersize=4, color="#E67E22", label="人数")
    axes[1].set_ylabel("数量")
    axes[1].legend(loc="upper left", ncol=2)

    axes[2].plot(x, _series(rows, 7), marker="o", linewidth=2.0, markersize=4, color="#34495E", label="整体人均")
    axes[2].plot(x, _series(rows, 9), marker="o", linewidth=1.8, markersize=3, color=THEME_BLUE, label="大厅")
    axes[2].plot(x, _series(rows, 11), marker="o", linewidth=1.8, markersize=3, color="#8E44AD", label="包房")
    axes[2].plot(x, _series(rows, 13), marker="o", linewidth=1.8, markersize=3, color="#C0392B", label="院子")
    axes[2].set_ylabel("人均")
    axes[2].legend(loc="upper left", ncol=4)

    tick_step = 2 if len(labels) > 14 else 1
    for ax in axes:
        ax.grid(True, linestyle="--", alpha=0.42, color="#cccccc")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.tick_params(axis="both", labelsize=8)
    axes[-1].set_xticks(x[::tick_step])
    axes[-1].set_xticklabels(labels[::tick_step], rotation=45, ha="right")
    fig.tight_layout(rect=[0, 0, 1, 0.96])

    path = output_dir / f"weekday_trend_{uuid.uuid4().hex}.png"
    fig.savefig(path)
    plt.close(fig)
    return path


def write_weekday_trend_pdf(config: WeekdayTrendConfig) -> Path:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    conn = _connect_readonly(config.db_path)
    try:
        rows_by_weekday = build_weekday_rows(conn, config)
    finally:
        conn.close()

    output_path = config.output_dir / f"周几半年趋势_{config.date_tag}_{config.store_tag}.pdf"
    styles = _pdf_styles()
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm,
    )
    story = [
        Paragraph("金谷仓周几半年趋势分析报告", styles["title"]),
        Paragraph(
            f"{config.store_tag} ｜ 统计周期：{config.first_monday:%Y-%m-%d} ~ "
            f"{(config.latest_monday + timedelta(days=6)):%Y-%m-%d}",
            styles["normal"],
        ),
        Spacer(1, 0.35*cm),
        Paragraph("一、最新一周按星期总览", styles["subtitle"]),
        _latest_summary_table(rows_by_weekday, styles),
        Spacer(1, 0.25*cm),
        Paragraph(
            "注：本报告营业额为堂食分桌有效消费团体口径，不含自取外卖、吧台及零食购买团体、第三方平台外卖；"
            "包房对应库内“包间”，院子对应库内“户外”。",
            styles["note"],
        ),
    ]

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        for index, weekday_name in enumerate(WEEKDAY_NAMES, start=1):
            story.append(PageBreak())
            story.append(Paragraph(f"{index + 1}、{weekday_name} 连续趋势", styles["subtitle"]))
            chart_path = _plot_weekday_chart(rows_by_weekday[weekday_name], weekday_name, tmp_dir, config.store_tag)
            story.append(Image(str(chart_path), width=16.5*cm, height=12.5*cm))
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(f"{weekday_name} 明细表", styles["normal"]))
            story.append(Spacer(1, 0.1*cm))
            story.append(_weekday_detail_table(rows_by_weekday[weekday_name], styles))
        doc.build(story)

    return output_path


def write_weekday_trend_reports(config: WeekdayTrendConfig) -> tuple[Path, Path]:
    return write_weekday_trend_workbook(config), write_weekday_trend_pdf(config)


def generate_weekday_trend_workbook(
    db_path: str | os.PathLike[str],
    output_dir: str | os.PathLike[str],
    end_date: str,
    store_name: str | None = None,
    weeks: int = 26,
) -> Path:
    if weeks < 2:
        raise ValueError("weeks 必须大于等于 2")
    config = WeekdayTrendConfig(
        db_path=Path(db_path),
        output_dir=Path(output_dir),
        end_date=_parse_date(end_date),
        store_name=store_name,
        weeks=weeks,
    )
    return write_weekday_trend_workbook(config)


def generate_weekday_trend_reports(
    db_path: str | os.PathLike[str],
    output_dir: str | os.PathLike[str],
    end_date: str,
    store_name: str | None = None,
    weeks: int = 26,
) -> tuple[Path, Path]:
    if weeks < 2:
        raise ValueError("weeks 必须大于等于 2")
    config = WeekdayTrendConfig(
        db_path=Path(db_path),
        output_dir=Path(output_dir),
        end_date=_parse_date(end_date),
        store_name=store_name,
        weeks=weeks,
    )
    return write_weekday_trend_reports(config)


def main() -> None:
    parser = argparse.ArgumentParser(description="生成周一至周日近半年趋势 Excel/PDF")
    parser.add_argument("--db", required=True, help="长期订单分析 SQLite 主库")
    parser.add_argument("--output-dir", default="./output", help="输出目录")
    parser.add_argument("--end-date", help="最新周期结束日期，格式 YYYY-MM-DD；不传则使用库内最新日期")
    parser.add_argument("--store", default=None, help="门店名称；不传则双店合计")
    parser.add_argument("--weeks", type=int, default=26, help="向前观察的自然周数，默认 26")
    args = parser.parse_args()

    db_path = Path(args.db)
    if args.end_date:
        end = _parse_date(args.end_date)
    else:
        conn = _connect_readonly(db_path)
        try:
            end = _latest_daily_date(conn, args.store)
        finally:
            conn.close()
    xlsx_path, pdf_path = write_weekday_trend_reports(
        WeekdayTrendConfig(
            db_path=db_path,
            output_dir=Path(args.output_dir),
            end_date=end,
            store_name=args.store,
            weeks=args.weeks,
        )
    )
    print(f"周几半年趋势 Excel 已生成: {xlsx_path}")
    print(f"周几半年趋势 PDF 已生成: {pdf_path}")


if __name__ == "__main__":
    main()
