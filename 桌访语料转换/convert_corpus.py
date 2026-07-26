#!/usr/bin/env python3
"""Convert 桌访数据 weekly CSV to 语料桌访_1.5版 xlsx (formatted like corpus template)."""

from __future__ import annotations

import argparse
import calendar
import re
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

OUTPUT_COLUMNS = [
    "会话ID",
    "服务员",
    "店面",
    "订单号",
    "桌台号",
    "就餐人数",
    "支付金额",
    "结账状态",
    "下单时间",
    "开单人",
    "会员状态",
    "会员手机号",
    "语音转录",
]

DEFAULT_TEMPLATE = Path(__file__).resolve().parent / "assets" / "语料桌访_1.5版_模板.xlsx"
DATE_RANGE_RE = re.compile(
    r"(\d{4})-(\d{1,2})-(\d{1,2})_至_(\d{4})-(\d{1,2})-(\d{1,2})"
)
FOLDER_BATCH_RE = re.compile(r"(?<!\d)(\d{6})(?!\d)")
CYCLE_RE = re.compile(r"(?<!\d)(\d{4})([WM])(\d{1,2})(?!\d)", re.IGNORECASE)


def _read_csv(path: Path) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return pd.read_csv(path, encoding=enc)
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"无法读取 CSV 编码: {path}")


def _parse_end_date_from_name(name: str) -> tuple[int, int, int] | None:
    m = DATE_RANGE_RE.search(name)
    if not m:
        return None
    return int(m.group(4)), int(m.group(5)), int(m.group(6))


def _parse_cycle_end_date(text: str) -> tuple[int, int, int] | None:
    m = CYCLE_RE.search(text)
    if not m:
        return None
    year = int(m.group(1))
    period_type = m.group(2).upper()
    number = int(m.group(3))
    if period_type == "W":
        end = date.fromisocalendar(year, number, 7)
    else:
        end = date(year, number, calendar.monthrange(year, number)[1])
    return end.year, end.month, end.day


def _detect_batch_prefix(csv_path: Path, end_date: tuple[int, int, int] | None) -> str:
    for part in reversed(csv_path.resolve().parts):
        m = FOLDER_BATCH_RE.fullmatch(part)
        if m:
            return m.group(1)
    cycle_text = " ".join(csv_path.resolve().parts[-6:]) + " " + csv_path.name
    cycle_end = _parse_cycle_end_date(cycle_text)
    if cycle_end:
        y, mo, d = cycle_end
        return f"{y % 100:02d}{mo:02d}{d:02d}"
    if end_date:
        y, mo, d = end_date
        return f"{y % 100:02d}{mo:02d}{d:02d}"
    return datetime.now().strftime("%y%m%d")


def _parse_order_time(value) -> tuple[int, datetime | None]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 1, None
    text = str(value).strip()
    if not text or text == "未识别":
        return 1, None
    try:
        return 0, pd.to_datetime(text)
    except Exception:
        return 1, None


def _build_output_df(df: pd.DataFrame, batch_prefix: str) -> pd.DataFrame:
    missing = [c for c in OUTPUT_COLUMNS if c != "会话ID" and c not in df.columns]
    if missing:
        raise ValueError(f"CSV 缺少必要列: {', '.join(missing)}")

    work = df.copy()
    work["_sort_rank"], work["_sort_dt"] = zip(*work["下单时间"].map(_parse_order_time))
    work = work.sort_values(["_sort_rank", "_sort_dt"], ascending=[True, True], kind="mergesort")
    work = work.reset_index(drop=True)
    work["会话ID"] = [f"{batch_prefix}-{i:03d}" for i in range(1, len(work) + 1)]

    out = work[OUTPUT_COLUMNS].copy()

    for col in ("就餐人数", "支付金额"):
        out[col] = pd.to_numeric(out[col], errors="coerce")

    def _to_excel_time(v):
        rank, dt = _parse_order_time(v)
        if rank == 1 or dt is None:
            return "未识别"
        return dt.to_pydatetime() if hasattr(dt, "to_pydatetime") else dt

    out["下单时间"] = out["下单时间"].map(_to_excel_time)

    for col in OUTPUT_COLUMNS:
        if col in ("就餐人数", "支付金额", "下单时间"):
            continue
        out[col] = out[col].fillna("").astype(str).str.strip()
        out[col] = out[col].replace({"nan": "", "None": ""})

    return out


def _output_filename(csv_path: Path, row_count: int, end_date: tuple[int, int, int] | None) -> str:
    if end_date:
        y, mo, d = end_date
        date_tag = f"{y}-{mo}-{d}"
    else:
        now = datetime.now()
        date_tag = f"{now.year}-{now.month}-{now.day}"
    return f"语料桌访_1.5版_{row_count}条_{date_tag}.xlsx"


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _estimate_row_height(text: str, base: float = 51.0) -> float:
    if not text:
        return base
    length = len(str(text))
    if length <= 80:
        return base
    if length <= 160:
        return 68.0
    if length <= 280:
        return 84.0
    if length <= 420:
        return 101.0
    return min(200.0, 118.0)


def _write_xlsx(
    df: pd.DataFrame,
    output_path: Path,
    template_path: Path,
) -> None:
    style_wb = load_workbook(template_path)
    style_ws = style_wb.active
    header_styles = [style_ws.cell(1, c) for c in range(1, len(OUTPUT_COLUMNS) + 1)]
    data_styles = [style_ws.cell(2, c) for c in range(1, len(OUTPUT_COLUMNS) + 1)]

    wb = load_workbook(template_path)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    ws.title = style_ws.title

    sheet_title = output_path.stem.replace("语料桌访", "桌探数据", 1)
    if len(sheet_title) <= 31:
        ws.title = sheet_title

    for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = style_ws.column_dimensions[letter].width

    ws.row_dimensions[1].height = style_ws.row_dimensions[1].height or 34.0
    for col_idx, header in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(1, col_idx, header)
        _copy_cell_style(header_styles[col_idx - 1], cell)

    time_col = OUTPUT_COLUMNS.index("下单时间") + 1
    transcript_col = OUTPUT_COLUMNS.index("语音转录") + 1

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        values = list(row)
        transcript = values[transcript_col - 1]
        ws.row_dimensions[row_idx].height = _estimate_row_height(transcript)

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx)
            _copy_cell_style(data_styles[col_idx - 1], cell)

            if col_idx == time_col:
                if value == "未识别" or value is None or value == "":
                    cell.value = "未识别"
                elif isinstance(value, datetime):
                    cell.value = value
                else:
                    try:
                        cell.value = pd.to_datetime(value).to_pydatetime()
                    except Exception:
                        cell.value = str(value)
            elif col_idx in (6, 7):
                if value == "" or value is None or (isinstance(value, float) and pd.isna(value)):
                    cell.value = None
                else:
                    cell.value = int(value) if float(value).is_integer() else float(value)
            else:
                cell.value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    style_wb.close()
    wb.close()


def convert_csv_to_corpus_xlsx(
    input_csv: Path,
    *,
    template: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    input_csv = input_csv.resolve()
    if not input_csv.exists():
        raise FileNotFoundError(input_csv)

    template_path = (template or DEFAULT_TEMPLATE).resolve()
    if not template_path.exists():
        raise FileNotFoundError(f"模板不存在: {template_path}")

    df_raw = _read_csv(input_csv)
    end_date = _parse_end_date_from_name(input_csv.name)
    batch_prefix = _detect_batch_prefix(input_csv, end_date)
    df_out = _build_output_df(df_raw, batch_prefix)

    if output_path is None:
        output_path = input_csv.parent / _output_filename(input_csv, len(df_out), end_date)
    else:
        output_path = output_path.resolve()

    _write_xlsx(df_out, output_path, template_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="将桌访数据 CSV 转换为语料桌访_1.5版 xlsx（格式对齐模板）"
    )
    parser.add_argument("input_csv", help="输入 CSV，如 桌访数据_2026-06-08_至_2026-06-14_467条.csv")
    parser.add_argument(
        "--template",
        default=str(DEFAULT_TEMPLATE),
        help="格式模板 xlsx（默认使用 assets/语料桌访_1.5版_模板.xlsx）",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出 xlsx 路径（默认与 CSV 同目录，自动命名）",
    )
    args = parser.parse_args()

    try:
        out = convert_csv_to_corpus_xlsx(
            Path(args.input_csv),
            template=Path(args.template),
            output_path=Path(args.output) if args.output else None,
        )
    except Exception as exc:
        print(f"错误: {exc}", file=sys.stderr)
        sys.exit(1)

    print(f"已生成: {out}")


if __name__ == "__main__":
    main()
