import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def write_excel_report(output_path, summary_data, store_comp_df, daily_trends_df, platform_stats_df, hourly_df, meal_df, eff_df, overtime_df, abnormal_df, detail_df):
    """
    Writes a beautifully formatted multi-sheet Excel report using openpyxl.
    """
    wb = Workbook()
    
    # Define styles
    font_family = "Microsoft YaHei"
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    section_font = Font(name=font_family, size=12, bold=True, color="333333")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="000000")
    total_font = Font(name=font_family, size=10, bold=True, color="000000")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='double', color='000000')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Helper to style a sheet
    def apply_standard_formatting(ws, start_row=4, is_abnormal=False):
        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row < start_row:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Sheet 1: 数据总览 (Data Overview)
    ws1 = wb.active
    ws1.title = "数据总览"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.append(["外卖平台经营数据总览"])
    ws1.cell(1, 1).font = title_font
    ws1.append([])
    
    # Summary Table Headers
    ws1.append(["指标项目", "合计值", "比例 / 均值"])
    for col_idx in range(1, 4):
        cell = ws1.cell(3, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    s_keys = [
        ("有效订单数", "{:,.0f}", None),
        ("退单数", "{:,.0f}", None),
        ("订单收入 (营业额)", "¥{:,.2f}", None),
        ("顾客实付", "¥{:,.2f}", None),
        ("客单价", "¥{:,.2f}", None),
        ("外卖抽佣", "¥{:,.2f}", None),
        ("抽佣率", None, "{:.2%}"),
        ("订单支出", "¥{:,.2f}", None),
        ("订单支出率", None, "{:.2%}"),
        ("部分退款", "¥{:,.2f}", None),
        ("订单金额", "¥{:,.2f}", None),
        ("菜品合计金额", "¥{:,.2f}", None),
        ("菜品收入", "¥{:,.2f}", None),
        ("菜品优惠", "¥{:,.2f}", None),
        ("订单优惠", "¥{:,.2f}", None),
        ("平台优惠", "¥{:,.2f}", None)
    ]
    
    curr_row = 4
    for idx, (label, val_fmt, pct_fmt) in enumerate(s_keys):
        val = summary_data.get(label.replace(" (营业额)", ""), 0.0)
        
        ws1.cell(curr_row, 1, label).alignment = align_left
        
        # Determine value & rate columns placement
        if val_fmt:
            ws1.cell(curr_row, 2, val).number_format = val_fmt.replace("{:,.0f}", "#,##0").replace("¥{:,.2f}", "¥#,##0.00")
            ws1.cell(curr_row, 3, "-").alignment = align_center
        else:
            ws1.cell(curr_row, 2, "-").alignment = align_center
            ws1.cell(curr_row, 3, val).number_format = "0.0%"
            
        for col_idx in range(1, 4):
            cell = ws1.cell(curr_row, col_idx)
            cell.font = data_font
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = alt_fill
                
        curr_row += 1
        
    apply_standard_formatting(ws1, start_row=3)
    
    # Sheet 2: 门店对比 (Store Comparison)
    ws2 = wb.create_sheet(title="门店对比")
    ws2.views.sheetView[0].showGridLines = True
    ws2.append(["各门店外卖经营对比"])
    ws2.cell(1, 1).font = title_font
    ws2.append([])
    
    # Write DataFrame
    for r in dataframe_to_rows(store_comp_df, index=False, header=True):
        ws2.append(r)
        
    # Style Header and rows
    ws2.row_dimensions[3].height = 24
    for col_idx in range(1, len(store_comp_df.columns) + 1):
        cell = ws2.cell(3, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(4, len(store_comp_df) + 4):
        ws2.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(store_comp_df.columns) + 1):
            cell = ws2.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format columns
            col_name = store_comp_df.columns[c_idx - 1]
            if "收入" in col_name or "实付" in col_name or "金额" in col_name or "客单价" in col_name:
                cell.number_format = "¥#,##0.00"
                cell.alignment = align_right
            elif "率" in col_name:
                cell.number_format = "0.0%"
                cell.alignment = align_right
            elif "订单数" in col_name or "退单" in col_name:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    apply_standard_formatting(ws2, start_row=3)
    
    # Sheet 3: 每日趋势 (Daily Trends)
    ws3 = wb.create_sheet(title="每日趋势")
    ws3.views.sheetView[0].showGridLines = True
    ws3.append(["每日外卖经营趋势"])
    ws3.cell(1, 1).font = title_font
    ws3.append([])
    
    for r in dataframe_to_rows(daily_trends_df, index=False, header=True):
        ws3.append(r)
        
    ws3.row_dimensions[3].height = 24
    for col_idx in range(1, len(daily_trends_df.columns) + 1):
        cell = ws3.cell(3, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(4, len(daily_trends_df) + 4):
        ws3.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(daily_trends_df.columns) + 1):
            cell = ws3.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = daily_trends_df.columns[c_idx - 1]
            if "收入" in col_name or "实付" in col_name or "金额" in col_name or "客单价" in col_name:
                cell.number_format = "¥#,##0.00"
                cell.alignment = align_right
            elif "率" in col_name:
                cell.number_format = "0.0%"
                cell.alignment = align_right
            elif "数" in col_name:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    apply_standard_formatting(ws3, start_row=3)
    
    # Sheet 4: 平台来源 (Platform Stats)
    ws4 = wb.create_sheet(title="平台来源")
    ws4.views.sheetView[0].showGridLines = True
    ws4.append(["平台来源占比分析"])
    ws4.cell(1, 1).font = title_font
    ws4.append([])
    
    for r in dataframe_to_rows(platform_stats_df, index=False, header=True):
        ws4.append(r)
        
    ws4.row_dimensions[3].height = 24
    for col_idx in range(1, len(platform_stats_df.columns) + 1):
        cell = ws4.cell(3, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(4, len(platform_stats_df) + 4):
        ws4.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(platform_stats_df.columns) + 1):
            cell = ws4.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = platform_stats_df.columns[c_idx - 1]
            if "收入" in col_name or "实付" in col_name or "金额" in col_name or "客单价" in col_name:
                cell.number_format = "¥#,##0.00"
                cell.alignment = align_right
            elif "率" in col_name:
                cell.number_format = "0.0%"
                cell.alignment = align_right
            elif "数" in col_name:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    apply_standard_formatting(ws4, start_row=3)
    
    # Sheet 5: 时段分布 (Time Distribution)
    ws5 = wb.create_sheet(title="时段分布")
    ws5.views.sheetView[0].showGridLines = True
    ws5.append(["外卖下单时段与午晚市分布"])
    ws5.cell(1, 1).font = title_font
    ws5.append([])
    
    # Section A: Meal Period Stats
    ws5.append(["一、午市 / 晚市占比"])
    ws5.cell(3, 1).font = section_font
    
    for r in dataframe_to_rows(meal_df, index=False, header=True):
        ws5.append(r)
        
    ws5.row_dimensions[4].height = 24
    for col_idx in range(1, len(meal_df.columns) + 1):
        cell = ws5.cell(4, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    meal_end = 5 + len(meal_df)
    for r_idx in range(5, meal_end):
        ws5.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(meal_df.columns) + 1):
            cell = ws5.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = meal_df.columns[c_idx - 1]
            if "收入" in col_name:
                cell.number_format = "¥#,##0.00"
                cell.alignment = align_right
            elif "占比" in col_name:
                cell.number_format = "0.0%"
                cell.alignment = align_right
            elif "订单数" in col_name:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    # Section B: Hourly Stats
    ws5.append([])
    ws5.append(["二、24小时下单时段细分"])
    ws5.cell(meal_end + 1, 1).font = section_font
    
    for r in dataframe_to_rows(hourly_df, index=False, header=True):
        ws5.append(r)
        
    h_header_row = meal_end + 2
    ws5.row_dimensions[h_header_row].height = 24
    for col_idx in range(1, len(hourly_df.columns) + 1):
        cell = ws5.cell(h_header_row, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(h_header_row + 1, h_header_row + len(hourly_df) + 1):
        ws5.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(hourly_df.columns) + 1):
            cell = ws5.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = hourly_df.columns[c_idx - 1]
            if "收入" in col_name:
                cell.number_format = "¥#,##0.00"
                cell.alignment = align_right
            elif "小时" in col_name or "订单数" in col_name:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    apply_standard_formatting(ws5, start_row=4)
    
    # Sheet 6: 履约效率 (Fulfillment Efficiency)
    ws6 = wb.create_sheet(title="履约效率")
    ws6.views.sheetView[0].showGridLines = True
    ws6.append(["履约时长与配送超时关注单"])
    ws6.cell(1, 1).font = title_font
    ws6.append([])
    
    # Section A: Fulfillment times
    ws6.append(["一、各门店履约时间统计 (分钟)"])
    ws6.cell(3, 1).font = section_font
    
    for r in dataframe_to_rows(eff_df, index=False, header=True):
        ws6.append(r)
        
    ws6.row_dimensions[4].height = 24
    for col_idx in range(1, len(eff_df.columns) + 1):
        cell = ws6.cell(4, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    eff_end = 5 + len(eff_df)
    for r_idx in range(5, eff_end):
        ws6.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(eff_df.columns) + 1):
            cell = ws6.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = eff_df.columns[c_idx - 1]
            if "门店" not in col_name:
                cell.number_format = "#,##0.0"
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    # Section B: Overtime alerts
    ws6.append([])
    ws6.append(["二、超时关注单 (送达时长超过45分钟)"])
    ws6.cell(eff_end + 1, 1).font = section_font
    
    for r in dataframe_to_rows(overtime_df, index=False, header=True):
        ws6.append(r)
        
    ot_header_row = eff_end + 2
    ws6.row_dimensions[ot_header_row].height = 24
    for col_idx in range(1, len(overtime_df.columns) + 1):
        cell = ws6.cell(ot_header_row, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(ot_header_row + 1, ot_header_row + len(overtime_df) + 1):
        ws6.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(overtime_df.columns) + 1):
            cell = ws6.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = overtime_df.columns[c_idx - 1]
            if "时长" in col_name:
                cell.number_format = "#,##0.0"
                cell.alignment = align_right
            elif "时间" in col_name:
                cell.alignment = align_center
            else:
                cell.alignment = align_center
                
    apply_standard_formatting(ws6, start_row=4)
    
    # Sheet 7: 退单异常 (Abnormal Exception)
    ws7 = wb.create_sheet(title="退单异常")
    ws7.views.sheetView[0].showGridLines = True
    ws7.append(["退单与经营数据异常列表"])
    ws7.cell(1, 1).font = title_font
    ws7.append([])
    
    for r in dataframe_to_rows(abnormal_df, index=False, header=True):
        ws7.append(r)
        
    ws7.row_dimensions[3].height = 24
    for col_idx in range(1, len(abnormal_df.columns) + 1):
        cell = ws7.cell(3, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(4, len(abnormal_df) + 4):
        ws7.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(abnormal_df.columns) + 1):
            cell = ws7.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = abnormal_df.columns[c_idx - 1]
            if "收入" in col_name:
                cell.number_format = "¥#,##0.00"
                cell.alignment = align_right
            elif "说明" in col_name or "异常" in col_name:
                cell.alignment = align_left
            elif "时间" in col_name:
                cell.alignment = align_center
            else:
                cell.alignment = align_center
                
    apply_standard_formatting(ws7, start_row=3)
    
    # Sheet 8: 脱敏明细 (Masked Details)
    ws8 = wb.create_sheet(title="脱敏明细")
    ws8.views.sheetView[0].showGridLines = True
    ws8.append(["外卖脱敏明细数据"])
    ws8.cell(1, 1).font = title_font
    ws8.append([])
    
    for r in dataframe_to_rows(detail_df, index=False, header=True):
        ws8.append(r)
        
    ws8.row_dimensions[3].height = 24
    for col_idx in range(1, len(detail_df.columns) + 1):
        cell = ws8.cell(3, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(4, len(detail_df) + 4):
        ws8.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(detail_df.columns) + 1):
            cell = ws8.cell(r_idx, c_idx)
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_fill
            # Format
            col_name = detail_df.columns[c_idx - 1]
            if col_name in ['订单金额', '菜品合计金额', '附加费分摊', '菜品优惠', '菜品收入', '餐盒费', '打包费', '配送费', '订单优惠', '平台优惠', '顾客实付', '订单支出', '外卖抽佣', '部分退款', '订单收入']:
                cell.number_format = "¥#,##0.00"
                cell.alignment = align_right
            elif "时间" in col_name or "对账" in col_name or "营业日" in col_name:
                cell.alignment = align_center
            elif "外卖订单号" in col_name or "收银" in col_name or "流水号" in col_name:
                cell.number_format = "@" # Force Text format
                cell.alignment = align_center
            else:
                cell.alignment = align_center
                
    apply_standard_formatting(ws8, start_row=3)
    
    # Save the file
    wb.save(output_path)
    print(f"Excel report successfully written to: {output_path}")
