"""
Excel writer for long-term order analysis.

Generates a 4-sheet workbook from daily statistics data.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _style_data_row(ws, row_num, num_cols, is_total=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = THIN_BORDER
        if is_total:
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00' if col > 2 else '@'
        else:
            if col == 1:
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0'
            elif col == 2:
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '@'
            elif col > 2 and ('百分比' in str(ws.cell(row=1, column=col).value or '') or '占比' in str(ws.cell(row=1, column=col).value or '')):
                cell.number_format = '0.0"%"'
            else:
                cell.number_format = '#,##0.00'


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or '')
                # Approximate: CJK chars count as 2
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, width)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def write_excel_report(all_dates, overview_rows, order_count_rows,
                       bucket_rows, opener_rows, all_openers, output_path):
    """
    Write the 4-sheet Excel report.

    Args:
        all_dates: sorted list of date strings
        overview_rows: list of (date, category, sub_category, 营业额, 百分比, 人数, 人均)
        order_count_rows: list of (date, 原始订单数, ...)
        bucket_rows: list of (date, bucket, 订单数, 占比)
        opener_rows: list of (date, opener_name, order_count, total_amount)
        all_openers: sorted list of unique opener names
        output_path: file path for the Excel output
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: 数据总览 ──
    ws1 = wb.active
    ws1.title = "数据总览"

    # Build per-date data dict for pivot
    # overview_rows is already in (date, category, sub_category, ...) format
    # We need to pivot this into wide format
    date_overview = {}
    for row in overview_rows:
        date, cat, sub, rev, pct, ppl, avg = row
        key = (cat, sub)
        if date not in date_overview:
            date_overview[date] = {}
        date_overview[date][key] = (rev, pct, int(ppl), avg)

    # Column headers for Sheet 1
    s1_headers = ['序号', '日期']
    s1_categories = [
        ('整体', ''), ('包间', ''), ('大厅', ''), ('户外', ''),
        ('午市', '整体'), ('午市', '包间'), ('午市', '大厅'), ('午市', '户外'),
        ('晚市', '整体'), ('晚市', '包间'), ('晚市', '大厅'), ('晚市', '户外'),
        ('会员', ''), ('非会员', ''),
    ]
    for cat, sub in s1_categories:
        label = f"{cat}{sub}" if sub else cat
        for suffix in ['营业额', '百分比', '人数', '人均']:
            s1_headers.append(f"{label}_{suffix}")

    for ci, h in enumerate(s1_headers, 1):
        ws1.cell(row=1, column=ci, value=h)
    _style_header(ws1, len(s1_headers))

    if all_dates:
        totals = {h: 0.0 for h in s1_headers[2:]}
        for si, date in enumerate(all_dates):
            row_num = si + 2
            ws1.cell(row=row_num, column=1, value=si + 1)
            ws1.cell(row=row_num, column=2, value=date)
            col = 3
            for cat, sub in s1_categories:
                key = (cat, sub)
                vals = date_overview.get(date, {}).get(key, (0, 0, 0, 0))
                for v in vals:
                    ws1.cell(row=row_num, column=col, value=v)
                    col += 1
            _style_data_row(ws1, row_num, len(s1_headers))

        # Totals row
        total_row = len(all_dates) + 2
        ws1.cell(row=total_row, column=1, value='')
        ws1.cell(row=total_row, column=2, value='合计')
        # Recompute totals across all dates
        col = 3
        for cat, sub in s1_categories:
            key = (cat, sub)
            total_rev = 0.0
            total_ppl = 0
            for date in all_dates:
                vals = date_overview.get(date, {}).get(key, (0, 0, 0, 0))
                total_rev += float(vals[0] or 0)
                total_ppl += int(vals[2] or 0)
            total_avg = round(total_rev / total_ppl, 2) if total_ppl > 0 else 0.0
            overall_rev = 0.0
            overall_ppl = 0
            for date in all_dates:
                ov = date_overview.get(date, {}).get(('整体', ''), (0, 0, 0, 0))
                overall_rev += float(ov[0] or 0)
                overall_ppl += int(ov[2] or 0)

            # percentage needs to handle different bases per category
            if cat in ('包间', '大厅', '户外', '会员', '非会员') and sub == '':
                total_pct = round(total_rev / overall_rev * 100, 1) if overall_rev > 0 else 0.0
            elif cat == '午市' and sub == '整体':
                total_pct = round(total_rev / overall_rev * 100, 1) if overall_rev > 0 else 0.0
            elif cat == '午市':
                wu_total = 0.0
                for date in all_dates:
                    wu_total += float(date_overview.get(date, {}).get(('午市', '整体'), (0, 0, 0, 0))[0] or 0)
                total_pct = round(total_rev / wu_total * 100, 1) if wu_total > 0 else 0.0
            elif cat == '晚市' and sub == '整体':
                total_pct = round(total_rev / overall_rev * 100, 1) if overall_rev > 0 else 0.0
            elif cat == '晚市':
                wan_total = 0.0
                for date in all_dates:
                    wan_total += float(date_overview.get(date, {}).get(('晚市', '整体'), (0, 0, 0, 0))[0] or 0)
                total_pct = round(total_rev / wan_total * 100, 1) if wan_total > 0 else 0.0
            else:
                total_pct = 100.0

            ws1.cell(row=total_row, column=col, value=round(total_rev, 2))
            ws1.cell(row=total_row, column=col + 1, value=total_pct)
            ws1.cell(row=total_row, column=col + 2, value=int(total_ppl))
            ws1.cell(row=total_row, column=col + 3, value=total_avg)
            col += 4
        _style_data_row(ws1, total_row, len(s1_headers), is_total=True)

    ws1.freeze_panes = 'A2'
    _auto_width(ws1)

    # ── Sheet 2: 订单数量明细 ──
    ws2 = wb.create_sheet("订单数量明细")
    s2_headers = ['序号', '日期', '原始订单数', '外卖订单数', '非堂食订单数', '免单订单数',
                  '被合并订单数', '合并后消费团体数', '零食团体数', '打包团体数', '零散小单团体数', '吧台团体数', '统计消费团体数']
    for ci, h in enumerate(s2_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header(ws2, len(s2_headers))

    date_counts = {r[0]: r[1:] for r in order_count_rows}
    if all_dates:
        totals = [0] * (len(s2_headers) - 2)
        for si, date in enumerate(all_dates):
            row_num = si + 2
            ws2.cell(row=row_num, column=1, value=si + 1)
            ws2.cell(row=row_num, column=2, value=date)
            vals = date_counts.get(date, (0,) * len(totals))
            for vi, v in enumerate(vals):
                ws2.cell(row=row_num, column=vi + 3, value=int(v or 0))
                totals[vi] += int(v or 0)
            _style_data_row(ws2, row_num, len(s2_headers))

        total_row = len(all_dates) + 2
        ws2.cell(row=total_row, column=1, value='')
        ws2.cell(row=total_row, column=2, value='合计')
        for vi, v in enumerate(totals):
            ws2.cell(row=total_row, column=vi + 3, value=v)
        _style_data_row(ws2, total_row, len(s2_headers), is_total=True)
    ws2.freeze_panes = 'A2'
    _auto_width(ws2)

    # ── Sheet 3: 客单价区间分布 ──
    ws3 = wb.create_sheet("客单价区间分布")
    s3_headers = ['序号', '日期',
                  '≥300_订单数', '≥300_占比',
                  '200~300_订单数', '200~300_占比',
                  '150~200_订单数', '150~200_占比',
                  '100~150_订单数', '100~150_占比',
                  '<100_订单数', '<100_占比']
    for ci, h in enumerate(s3_headers, 1):
        ws3.cell(row=1, column=ci, value=h)
    _style_header(ws3, len(s3_headers))

    date_buckets = {}
    for r in bucket_rows:
        date, bucket, cnt, pct = r
        if date not in date_buckets:
            date_buckets[date] = {}
        date_buckets[date][bucket] = (int(cnt), float(pct))

    bucket_order = ['≥300', '200~300', '150~200', '100~150', '<100']
    if all_dates:
        totals_b = {b: [0, 0.0] for b in bucket_order}
        for si, date in enumerate(all_dates):
            row_num = si + 2
            ws3.cell(row=row_num, column=1, value=si + 1)
            ws3.cell(row=row_num, column=2, value=date)
            col = 3
            for bk in bucket_order:
                cnt, pct = date_buckets.get(date, {}).get(bk, (0, 0.0))
                ws3.cell(row=row_num, column=col, value=cnt)
                ws3.cell(row=row_num, column=col + 1, value=pct)
                totals_b[bk][0] += cnt
                totals_b[bk][1] = 0.0  # recalc below
                col += 2
            _style_data_row(ws3, row_num, len(s3_headers))

        # Recalc percentages for totals
        total_all = sum(totals_b[b][0] for b in bucket_order)
        for bk in bucket_order:
            totals_b[bk][1] = round(totals_b[bk][0] / total_all * 100, 1) if total_all > 0 else 0.0

        total_row = len(all_dates) + 2
        ws3.cell(row=total_row, column=1, value='')
        ws3.cell(row=total_row, column=2, value='合计')
        col = 3
        for bk in bucket_order:
            ws3.cell(row=total_row, column=col, value=totals_b[bk][0])
            ws3.cell(row=total_row, column=col + 1, value=totals_b[bk][1])
            col += 2
        _style_data_row(ws3, total_row, len(s3_headers), is_total=True)
    ws3.freeze_panes = 'A2'
    _auto_width(ws3)

    # ── Sheet 4: 开单人统计 ──
    ws4 = wb.create_sheet("开单人统计")
    s4_headers = ['序号', '日期']
    for op in all_openers:
        display = '扫码点餐' if op == '顾客/系统' else op
        s4_headers.append(f"{display}_数量")
        s4_headers.append(f"{display}_金额")

    for ci, h in enumerate(s4_headers, 1):
        ws4.cell(row=1, column=ci, value=h)
    _style_header(ws4, len(s4_headers))

    date_openers = {}
    for r in opener_rows:
        date, op_name, cnt, amt = r
        if date not in date_openers:
            date_openers[date] = {}
        date_openers[date][op_name] = (int(cnt), float(amt))

    if all_dates and all_openers:
        totals_o = {op: [0, 0.0] for op in all_openers}
        for si, date in enumerate(all_dates):
            row_num = si + 2
            ws4.cell(row=row_num, column=1, value=si + 1)
            ws4.cell(row=row_num, column=2, value=date)
            col = 3
            for op in all_openers:
                cnt, amt = date_openers.get(date, {}).get(op, (0, 0.0))
                ws4.cell(row=row_num, column=col, value=cnt)
                ws4.cell(row=row_num, column=col + 1, value=amt)
                totals_o[op][0] += cnt
                totals_o[op][1] += amt
                col += 2
            _style_data_row(ws4, row_num, len(s4_headers))

        total_row = len(all_dates) + 2
        ws4.cell(row=total_row, column=1, value='')
        ws4.cell(row=total_row, column=2, value='合计')
        col = 3
        for op in all_openers:
            ws4.cell(row=total_row, column=col, value=totals_o[op][0])
            ws4.cell(row=total_row, column=col + 1, value=round(totals_o[op][1], 2))
            col += 2
        _style_data_row(ws4, total_row, len(s4_headers), is_total=True)
    ws4.freeze_panes = 'A2'
    _auto_width(ws4)

    wb.save(output_path)
    print(f"Excel 报告已生成: {output_path}")
