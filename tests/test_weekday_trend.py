import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "周期对比分析"))

from weekday_trend import (  # noqa: E402
    WEEKDAY_NAMES,
    WeekdayTrendConfig,
    build_weekday_rows,
    write_weekday_trend_pdf,
    write_weekday_trend_workbook,
)


def _make_db(path: Path) -> None:
    conn = sqlite3.connect(path)
    conn.executescript(
        """
        CREATE TABLE daily_overview (
            date TEXT NOT NULL,
            store_name TEXT NOT NULL,
            category TEXT NOT NULL,
            sub_category TEXT DEFAULT '',
            营业额 REAL DEFAULT 0,
            百分比 REAL DEFAULT 0,
            人数 REAL DEFAULT 0,
            人均 REAL DEFAULT 0,
            PRIMARY KEY (date, store_name, category, sub_category)
        );
        CREATE TABLE daily_order_counts (
            date TEXT NOT NULL,
            store_name TEXT NOT NULL,
            原始订单数 INTEGER DEFAULT 0,
            外卖订单数 INTEGER DEFAULT 0,
            非堂食订单数 INTEGER DEFAULT 0,
            免单订单数 INTEGER DEFAULT 0,
            被合并订单数 INTEGER DEFAULT 0,
            合并后消费团体数 INTEGER DEFAULT 0,
            零食团体数 INTEGER DEFAULT 0,
            打包团体数 INTEGER DEFAULT 0,
            零散小单团体数 INTEGER DEFAULT 0,
            吧台团体数 INTEGER DEFAULT 0,
            统计消费团体数 INTEGER DEFAULT 0,
            PRIMARY KEY (date, store_name)
        );
        """
    )
    overview_rows = [
        ("2026-06-22", "万荷店", "整体", "", 1000, 0, 10, 100),
        ("2026-06-22", "万荷店", "大厅", "", 400, 0, 5, 80),
        ("2026-06-22", "万荷店", "包间", "", 450, 0, 3, 150),
        ("2026-06-22", "万荷店", "户外", "", 150, 0, 2, 75),
        ("2026-06-29", "万荷店", "整体", "", 1200, 0, 12, 100),
        ("2026-06-29", "万荷店", "大厅", "", 600, 0, 6, 100),
        ("2026-06-29", "万荷店", "包间", "", 400, 0, 4, 100),
        ("2026-06-29", "万荷店", "户外", "", 200, 0, 2, 100),
    ]
    conn.executemany(
        """
        INSERT INTO daily_overview
        (date, store_name, category, sub_category, 营业额, 百分比, 人数, 人均)
        VALUES (?,?,?,?,?,?,?,?)
        """,
        overview_rows,
    )
    conn.executemany(
        """
        INSERT INTO daily_order_counts
        (date, store_name, 统计消费团体数)
        VALUES (?,?,?)
        """,
        [
            ("2026-06-22", "万荷店", 8),
            ("2026-06-29", "万荷店", 9),
        ],
    )
    conn.commit()
    conn.close()


def test_build_weekday_rows_uses_dining_group_metrics(tmp_path):
    db_path = tmp_path / "trend.db"
    _make_db(db_path)
    config = WeekdayTrendConfig(
        db_path=db_path,
        output_dir=tmp_path,
        end_date=__import__("datetime").date(2026, 7, 5),
        store_name="万荷店",
        weeks=2,
    )
    conn = sqlite3.connect(db_path)
    try:
        rows = build_weekday_rows(conn, config)
    finally:
        conn.close()

    monday_rows = rows["周一"]
    assert len(monday_rows) == 2
    assert monday_rows[0][4] == 1000
    assert monday_rows[0][5] == 8
    assert monday_rows[0][6] == 10
    assert monday_rows[0][7] == 100
    assert monday_rows[0][8:14] == [5, 80, 3, 150, 2, 75]
    assert monday_rows[1][4] == 1200
    assert monday_rows[1][5] == 9


def test_write_weekday_trend_workbook_has_seven_weekday_sheets(tmp_path):
    db_path = tmp_path / "trend.db"
    _make_db(db_path)
    path = write_weekday_trend_workbook(
        WeekdayTrendConfig(
            db_path=db_path,
            output_dir=tmp_path,
            end_date=__import__("datetime").date(2026, 7, 5),
            store_name="万荷店",
            weeks=2,
        )
    )

    workbook = load_workbook(path, read_only=False)
    assert workbook.sheetnames == WEEKDAY_NAMES
    monday = workbook["周一"]
    assert monday["E5"].value == 1000
    assert monday["F5"].value == 8
    assert len(monday._charts) == 3


def test_write_weekday_trend_pdf_is_readable(tmp_path):
    db_path = tmp_path / "trend.db"
    _make_db(db_path)
    path = write_weekday_trend_pdf(
        WeekdayTrendConfig(
            db_path=db_path,
            output_dir=tmp_path,
            end_date=__import__("datetime").date(2026, 7, 5),
            store_name="万荷店",
            weeks=2,
        )
    )

    reader = PdfReader(str(path))
    assert len(reader.pages) >= 8
